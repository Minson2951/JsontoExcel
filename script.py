import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import AutoFilter

# Load JSON data
json_path  = r"E:\Python_test\json_to_xlsx\personaldetails.json"
df = pd.read_json(json_path)

# Save DataFrame to Excel
excel_path = r'E:\Python_test\json_to_xlsx\data.xlsx'
df.to_excel(excel_path, index=False, engine='openpyxl')

# Load the workbook and select the active worksheet
wb = load_workbook(excel_path)
ws = wb.active

# Apply formatting
top_row_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
for cell in ws[1]:
    cell.fill = top_row_fill

# Autofit columns (approximate)
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Enable filter
ws.auto_filter.ref = ws.dimensions

# Save the workbook
wb.save(excel_path)

print(f"Excel file saved as {excel_path}")
